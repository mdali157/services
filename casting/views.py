import base64
import datetime
import io
import uuid
from io import BytesIO

import openpyxl
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.db.models import Max, Sum
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.template.loader import get_template
from django.urls import reverse
from openpyxl.utils import get_column_letter
from weasyprint import HTML

from customers.models import Customer
from receiving.models import ServiceType, Receiving

from .models import *


# Create your views here.
@login_required
def add_casting(request):
    if request.method == 'POST':
        # 1. Read form fields
        customer_name = request.POST.get("customer_name", "").strip().capitalize()
        customer_phone = request.POST.get("customer_phone", "").strip()
        karate_value = request.POST.get("karate", "").strip().capitalize()
        color_value = request.POST.get("color", "").strip().capitalize()
        description = request.POST.get("description", "").strip()
        creation_date = request.POST.get("creation_date")  # YYYY-MM-DD
        captured_image_data = request.POST.get("captured_image")  # data:image/png;base64,....

        # 2. Get or create customer
        customer, created = Customer.objects.get_or_create(
            name=customer_name,
            phone_number=customer_phone,
            defaults={'added_by': request.user}
        )

        karate, created = Karate.objects.get_or_create(
            name=karate_value
        )

        color, created = Color.objects.get_or_create(
            name=color_value
        )

        # 3. Build Casting instance
        casting = Casting(
            customer=customer,
            karate=karate.name,
            color=color.name,
            description=description,
            created_at=creation_date,
            created_by=request.user,
            modified_by=request.user,
        )

        # 4. Handle base64 image
        if captured_image_data and ';base64,' in captured_image_data:
            header, b64data = captured_image_data.split(';base64,')
            ext = header.split('/')[-1]  # e.g. "png" or "jpeg"
            binary = base64.b64decode(b64data)
            fname = f"casting_{uuid.uuid4().hex[:10]}.{ext}"
            img_file = InMemoryUploadedFile(
                BytesIO(binary),
                field_name='image',
                name=fname,
                content_type=f"image/{ext}",
                size=len(binary),
                charset=None
            )
            casting.image = img_file

        # 5. Save (casting_no auto-assigned in model.save())
        casting.save()

        # 6. Respond with JSON for your AJAX
        return JsonResponse({
            'print_url': reverse('casting:print_casting_slip', args=[casting.id]),
            'redirect_url': reverse('casting:add_casting')
        })

    # GET: figure next casting_no and render
    max_no = Casting.objects.aggregate(max=Max('casting_no'))['max']
    next_no = 111111 if max_no is None else max_no + 1

    return render(request, 'casting/add_casting.html', {
        'next_casting_no': next_no
    })


@login_required
def update_casting(request, casting_id):
    casting = get_object_or_404(Casting, pk=casting_id)

    if request.method == 'POST':
        # 1. Read form fields
        karate_value = request.POST.get("karate", "").strip().capitalize()
        color_value = request.POST.get("color", "").strip().capitalize()
        description = request.POST.get("description", "").strip()
        delivery_remarks = request.POST.get("delivery_remarks", "").strip()
        creation_date = request.POST.get("creation_date")  # YYYY-MM-DD
        captured_image_data = request.POST.get("captured_image")  # data:image/png;base64,…
        is_delivered = request.POST.get('is_delivered')
        cash_received = request.POST.get('cash_received')
        gold_received = request.POST.get('gold_received')

        karate_obj, _ = Karate.objects.get_or_create(name=karate_value)
        color_obj, _ = Color.objects.get_or_create(name=color_value)

        # 3. Update fields on casting
        casting.karate = karate_obj.name
        casting.color = color_obj.name
        casting.description = description
        casting.remarks = delivery_remarks
        casting.created_at = creation_date
        casting.gold_received = gold_received
        casting.cash_received = cash_received
        casting.is_delivered = is_delivered == 'on'
        casting.modified_by = request.user

        # 4. Handle new base64 image if provided
        if captured_image_data and ';base64,' in captured_image_data:
            header, b64data = captured_image_data.split(';base64,')
            ext = header.split('/')[-1]
            binary = base64.b64decode(b64data)
            fname = f"casting_{uuid.uuid4().hex[:10]}.{ext}"
            img_file = InMemoryUploadedFile(
                BytesIO(binary),
                field_name='image',
                name=fname,
                content_type=f"image/{ext}",
                size=len(binary),
                charset=None
            )
            casting.delivery_image = img_file

        casting.save()
        return redirect('casting:all_casting')

    return render(request, 'casting/update_casting.html', {
        'casting': casting
    })


def search_karate(request):
    query = request.GET.get('q', '')
    all_karate = Karate.objects.filter(name__icontains=query)
    karate_suggestions = [
        {'name': karate.name}
        for karate in all_karate
    ]
    return JsonResponse({'suggestions': karate_suggestions})


def search_casting(request):
    query = request.GET.get('q', '')
    all_colors = Color.objects.filter(name__icontains=query)
    color_suggestions = [
        {'name': color.name}
        for color in all_colors
    ]
    return JsonResponse({'suggestions': color_suggestions})


def print_casting_slip(request, casting_id):
    casting = get_object_or_404(Casting, pk=casting_id)
    return render(request, 'casting/casting_print_slip.html', {'casting': casting})


def casting_flask_print_slip(request, casting_id):
    casting = get_object_or_404(Casting, pk=casting_id)
    service_charges_amount = float(casting.service_charges_amount) if casting.service_charges_amount is not None else 0
    cash_received = float(casting.cash_received) if casting.cash_received is not None else 0
    total_balance = service_charges_amount - cash_received
    total_weight24k = float(casting.total_weight24k) if casting.total_weight24k is not None else 0
    gold_received = float(casting.gold_received) if casting.gold_received is not None else 0

    total_gold_balance = total_weight24k - gold_received

    return render(request, 'casting/casting_flask_print_slip.html', {'casting': casting, 'total_balance': total_balance, 'total_gold_balance': total_gold_balance})


def all_flask(request):
    flasks = Flask.objects.all().order_by('-created_at')
    return render(request, 'casting/all_flask.html', {
        'all_flask': flasks,
    })


def all_casting(request):
    status = request.GET.get('status', 'undelivered')

    if status == 'delivered':
        all_casting = Casting.objects.filter(is_delivered=True)
    elif status == 'undelivered':
        all_casting = Casting.objects.filter(is_delivered=False)
    else:  # 'all'
        all_casting = Casting.objects.all()

    return render(request, 'casting/all_casting.html', {
        'castings_with_flask': all_casting,
        'selected_status': status,
    })


@login_required
def add_flask(request):
    # 1. Determine next_flask_id (just max ID + 1)
    next_id = (Flask.objects.aggregate(Max('id'))['id__max'] or 0) + 1

    # 2. Fetch all castings not yet assigned to a flask
    unflasked = Casting.objects.filter(flask__isnull=True)

    if request.method == 'POST':
        # 3. Extract and normalize karate & color strings
        karate_value = request.POST.get("karate", "").strip().capitalize()
        color_value = request.POST.get("color", "").strip().capitalize()

        # 4. Get or create the Karate and Color objects
        karate_obj, _ = Karate.objects.get_or_create(name=karate_value)
        color_obj, _ = Color.objects.get_or_create(name=color_value)

        # 5. Create the Flask instance (saving karate_obj.name & color_obj.name, to mirror how Casting saved them)
        flask = Flask.objects.create(
            karate=karate_obj.name,
            color=color_obj.name,
            input_weight=request.POST.get("input_weight", ""),
            output_weight=request.POST.get("output_weight", ""),
            machine_wastage=request.POST.get("machine_wastage", ""),
            production_weight=request.POST.get("production_weight", ""),
            created_at=request.POST.get("creation_date", ""),

        )

        # 6. Loop through POST keys to find all selected castings (those ending with "_sno")
        for key in request.POST:
            if key.startswith("casting_") and key.endswith("_sno"):
                base = key.replace("_sno", "")  # e.g. "casting_12"
                casting_id = base.split("_")[1]
                try:
                    casting = Casting.objects.get(id=casting_id)
                except Casting.DoesNotExist:
                    continue

                # 7. Update each Casting’s fields
                casting.weight = request.POST.get(f"{base}_weight")
                casting.converted24k = request.POST.get(f"{base}_converted")
                casting.total_weight24k = request.POST.get(f"{base}_total")
                casting.wastage_percentage = request.POST.get(f"{base}_wastage_percent")
                casting.wastage_weight = request.POST.get(f"{base}_wastage_weight")
                casting.service_charges_rate = request.POST.get(f"{base}_service_rate")
                casting.service_charges_amount = request.POST.get(f"{base}_service_amount")
                casting.flask = flask
                casting.modified_by = request.user
                casting.save()

        # 8. Redirect back to the “all_flask” page
        return redirect('casting:all_flask')

    # For GET requests, render the form with next_flask_id & unflasked_castings
    return render(request, 'casting/add_flask.html', {
        'next_flask_id': next_id,
        'unflasked_castings': unflasked,
    })


@login_required
def update_flask(request, flask_id):
    flask = get_object_or_404(Flask, id=flask_id)

    # 1. All castings already linked to this flask
    existing_castings = Casting.objects.filter(flask=flask)

    # 2. Castings not yet linked to any flask (for the “Add New Castings” select)
    unlinked_castings = Casting.objects.filter(flask__isnull=True)

    if request.method == 'POST':
        # a) Normalize & get_or_create Karate and Color
        karate_value = request.POST.get("karate", "").strip().capitalize()
        color_value = request.POST.get("color", "").strip().capitalize()

        karate_obj, _ = Karate.objects.get_or_create(name=karate_value)
        color_obj, _ = Color.objects.get_or_create(name=color_value)

        # b) Update Flask fields
        flask.karate = karate_obj.name
        flask.color = color_obj.name
        flask.input_weight = request.POST.get("input_weight", "")
        flask.output_weight = request.POST.get("output_weight", "")
        flask.machine_wastage = request.POST.get("machine_wastage", "")
        flask.production_weight = request.POST.get("production_weight", "")
        flask.created_at = request.POST.get("creation_date", "")
        flask.save()

        # c) Update data for castings already linked
        for casting in existing_castings:
            prefix = f"casting_{casting.id}_"
            casting.weight = request.POST.get(f"{prefix}weight")
            casting.converted24k = request.POST.get(f"{prefix}converted")
            casting.total_weight24k = request.POST.get(f"{prefix}total")
            casting.wastage_percentage = request.POST.get(f"{prefix}wastage_percent")
            casting.wastage_weight = request.POST.get(f"{prefix}wastage_weight")
            casting.service_charges_rate = request.POST.get(f"{prefix}service_rate")
            casting.service_charges_amount = request.POST.get(f"{prefix}service_amount")
            casting.modified_by = request.user
            casting.save()

        # d) Link & initialize data for any newly selected castings
        selected_ids = request.POST.getlist("related_castings")
        for casting_id in selected_ids:
            # Only link if it wasn't already linked
            if not existing_castings.filter(id=casting_id).exists():
                casting = get_object_or_404(Casting, id=casting_id)
                prefix = f"casting_{casting.id}_"
                casting.weight = request.POST.get(f"{prefix}weight")
                casting.converted24k = request.POST.get(f"{prefix}converted")
                casting.total_weight24k = request.POST.get(f"{prefix}total")
                casting.wastage_percentage = request.POST.get(f"{prefix}wastage_percent")
                casting.wastage_weight = request.POST.get(f"{prefix}wastage_weight")
                casting.service_charges_rate = request.POST.get(f"{prefix}service_rate")
                casting.service_charges_amount = request.POST.get(f"{prefix}service_amount")
                casting.flask = flask
                casting.modified_by = request.user
                casting.save()

        return redirect('casting:all_flask')

    # GET → render template with context
    return render(request, 'casting/update_flask.html', {
        'flask': flask,
        'existing_castings': existing_castings,
        'unlinked_castings': unlinked_castings,
    })


def report(request):
    service_type_list = ServiceType.objects.all()
    customers = Customer.objects.all()

    if request.method == "POST":
        # 1) Read form fields
        report_for = request.POST.get("report_for")
        from_date_str = request.POST.get("from_date")
        to_date_str = request.POST.get("to_date")
        report_type = request.POST.get("report_type")
        selected_customers = request.POST.getlist("customers")
        selected_service_type = request.POST.get("service_type")
        action = request.POST.get("action")

        # 2) Parse dates into date objects
        from_date = datetime.datetime.strptime(from_date_str, "%Y-%m-%d").date()
        to_date = datetime.datetime.strptime(to_date_str, "%Y-%m-%d").date()

        # 3) Build queryset depending on `report_for`
        if report_for == "Services":
            if report_type == "Summary":
                qs = Receiving.objects.filter(created_at__range=(from_date, to_date))
                if selected_customers:
                    qs = qs.filter(customer__id__in=selected_customers)
                if selected_service_type and selected_service_type != "All":
                    qs = qs.filter(service_type=selected_service_type)

                # Group by date and service type, and sum Estimate and Amount
                qs = qs.values('created_at', 'service_type') \
                    .annotate(
                    total_estimate=Sum('estimated_price'),
                    total_amount=Sum('actual_price')
                ).order_by('created_at', 'service_type')

                # Calculate totals
                totals = {
                    'grand_total_estimate': sum(item['total_estimate'] for item in qs if item['total_estimate']),
                    'grand_total_amount': sum(item['total_amount'] for item in qs if item['total_amount'])
                }
            else:  # Detail
                qs = Receiving.objects.filter(created_at__range=(from_date, to_date))
                if selected_customers:
                    qs = qs.filter(customer__id__in=selected_customers)
                if selected_service_type and selected_service_type != "All":
                    qs = qs.filter(service_type=selected_service_type)

                totals = {
                    'grand_total_estimate': sum(item.estimated_price for item in qs if item.estimated_price),
                    'grand_total_amount': sum(item.actual_price for item in qs if item.actual_price)
                }

        else:  # report_for == "Casting"
            if report_type == "Summary":
                qs = Casting.objects.filter(created_at__range=(from_date, to_date))
                if selected_customers:
                    qs = qs.filter(customer__id__in=selected_customers)

                # Group by date and sum other fields
                qs = qs.values('created_at') \
                    .annotate(
                    total_input_weight=Sum('flask__input_weight'),
                    total_output_weight=Sum('flask__output_weight'),
                    total_machine_wastage=Sum('flask__machine_wastage'),
                    total_production_weight=Sum('flask__production_weight'),
                    total_weight24k=Sum('total_weight24k'),
                    total_wastage_weight=Sum('wastage_weight'),
                    total_gold_received=Sum('gold_received'),
                    total_service_charges_amount=Sum('service_charges_amount'),
                    total_cash_received=Sum('cash_received')
                ).order_by('created_at')

                for item in qs:
                    item["balance_cash"] = float(item.get("total_service_charges_amount") or 0) - float(
                        item.get("total_cash_received") or 0)
                    item["balance_gold"] = float(item.get("total_weight24k") or 0) - float(
                        item.get("total_gold_received") or 0)

                totals = {
                    'grand_total_input': sum(
                        float(item['total_input_weight']) for item in qs if item['total_input_weight']),
                    'grand_total_output': sum(
                        float(item['total_output_weight']) for item in qs if item['total_output_weight']),
                    'grand_total_wastage': sum(
                        float(item['total_machine_wastage']) for item in qs if item['total_machine_wastage']),
                    'grand_total_production': sum(
                        float(item['total_production_weight']) for item in qs if item['total_production_weight']),
                    'grand_total_24k': sum(float(item['total_weight24k']) for item in qs if item['total_weight24k']),
                    'grand_total_wastage_weight': sum(
                        float(item['total_wastage_weight']) for item in qs if item['total_wastage_weight']),
                    'grand_total_gold': sum(
                        float(item['total_gold_received']) for item in qs if item['total_gold_received']),
                    'grand_total_service': sum(float(item['total_service_charges_amount']) for item in qs if
                                               item['total_service_charges_amount']),
                    'grand_total_cash': sum(
                        float(item['total_cash_received']) for item in qs if item['total_cash_received']),
                    'grand_total_balance_cash': sum(item['balance_cash'] for item in qs),
                    'grand_total_balance_gold': sum(item['balance_gold'] for item in qs),

                }
            else:  # Detail
                qs = Casting.objects.filter(created_at__range=(from_date, to_date))

                if selected_customers:
                    qs = qs.filter(customer__id__in=selected_customers)

                for item in qs:
                    item.balance_gold = (float(item.total_weight24k or 0) - float(item.gold_received or 0))
                    item.balance_cash = (float(item.service_charges_amount or 0) - float(item.cash_received or 0))

                totals = {
                    'grand_total_input': sum(
                        float(item.flask.input_weight) for item in qs if item.flask and item.flask.input_weight),
                    'grand_total_output': sum(
                        float(item.flask.output_weight) for item in qs if item.flask and item.flask.output_weight),
                    'grand_total_wastage': sum(
                        float(item.flask.machine_wastage) for item in qs if item.flask and item.flask.machine_wastage),
                    'grand_total_production': sum(float(item.flask.production_weight) for item in qs if
                                                  item.flask and item.flask.production_weight),
                    'grand_total_24k': sum(float(item.total_weight24k) for item in qs if item.total_weight24k),
                    'grand_total_wastage_weight': sum(float(item.wastage_weight) for item in qs if item.wastage_weight),
                    'grand_total_gold': sum(float(item.gold_received) for item in qs if item.gold_received),
                    'grand_total_service': sum(
                        float(item.service_charges_amount) for item in qs if item.service_charges_amount),
                    'grand_total_cash': sum(float(item.cash_received) for item in qs if item.cash_received),
                    'grand_total_balance_cash': sum(item.balance_cash for item in qs),
                    'grand_total_balance_gold': sum(item.balance_gold for item in qs)

                }

        # 4) Branch on action
        if action == "pdf":
            return generate_pdf_response(
                request=request,
                queryset=qs,
                report_for=report_for,
                report_type=report_type,
                from_date=from_date,
                to_date=to_date,
                totals=totals
            )
        else:  # action == "excel"
            return generate_excel_response(
                queryset=qs,
                report_for=report_for,
                report_type=report_type,
                from_date=from_date,
                to_date=to_date,
                totals=totals
            )

    # If GET, just render the filter form
    return render(
        request,
        "casting/reports.html",
        {"service_type": service_type_list, "customers": customers},
    )


def generate_pdf_response(request, queryset, report_for, report_type, from_date, to_date, totals):
    context = {
        "report_for": report_for,
        "report_type": report_type,
        "from_date": from_date,
        "to_date": to_date,
        "items": queryset,
        "totals": totals
    }

    template = get_template("casting/report_pdf.html")
    html_string = template.render(context)

    # Render to PDF using WeasyPrint
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="report.pdf"'
    return response


def generate_excel_response(queryset, report_for, report_type, from_date, to_date, totals):
    wb = openpyxl.Workbook()
    ws = wb.active

    if report_for == "Services":
        if report_type == "Detail":
            headers = [
                "Service #",
                "Date",
                "Type",
                "Customer Name",
                "Mobile #",
                "Work Description",
                "Estimate",
                "Amount",
                "Remarks",
            ]
        else:  # Summary
            headers = ["Date", "Type", "Estimate", "Amount"]
    else:  # report_for == "Casting"
        if report_type == "Detail":
            headers = [
                "Flask #", "Date", "Karate", "Color", "Input WT", "Output WT", "Machine Wastage",
                "Cast #", "Party Name", "Production Weight", "Weight-24K", "Wastage %",
                "Wastage Wt", "Total Wt", "Gold Received", "Balance Gold","Service Charges Rate", "Amount",
                "Received Amount", "Balance Cash", "Remarks"
            ]
        else:  # Summary
            headers = [
                "Date",
                "Input WT",
                "Output WT",
                "Machine Wastage",
                "Production Weight",
                "Weight-24K",
                "Wastage Wt",
                "Total Wt",
                "Gold Received",
                "Balance Gold",
                "Service Amount",
                "Received Amount",
                "Balance Cash"
            ]

    for col_idx, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = column_title
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(column_title) * 1.2, 15)

    # Write each row
    row_num = 2
    for obj in queryset:
        row = []
        if report_for == "Services":
            if report_type == "Detail":
                row = [
                    obj.service_no,
                    obj.created_at.strftime("%Y-%m-%d"),
                    obj.service_type,
                    obj.customer.name,
                    obj.customer.phone_number,
                    obj.description,
                    obj.estimated_price or "",
                    obj.actual_price or "",
                    obj.remarks or "",
                ]
            else:  # Summary
                row = [
                    obj["created_at"].strftime("%Y-%m-%d"),
                    obj["service_type"],
                    obj["total_estimate"] or "",
                    obj["total_amount"] or "",
                ]
        else:  # report_for == "Casting"
            if report_type == "Detail":
                row = [
                    obj.flask.id if obj.flask else "",
                    obj.created_at.strftime("%Y-%m-%d"),
                    obj.karate or "",
                    obj.color or "",
                    obj.flask.input_weight if obj.flask else "",
                    obj.flask.output_weight if obj.flask else "",
                    obj.flask.machine_wastage if obj.flask else "",
                    obj.casting_no,
                    obj.customer.name if obj.customer.name else "",
                    obj.flask.production_weight if obj.flask else "",
                    obj.total_weight24k or "",
                    obj.wastage_percentage or "",
                    obj.wastage_weight or "",
                    obj.total_weight24k or "",
                    obj.gold_received or "",
                    obj.balance_gold,
                    obj.service_charges_rate or "",
                    obj.service_charges_amount or "",
                    obj.cash_received or "",
                    obj.balance_cash,
                    obj.remarks if obj.remarks else "",
                ]
            else:  # Summary
                row = [
                    obj["created_at"].strftime("%Y-%m-%d"),
                    obj["total_input_weight"] or "",
                    obj["total_output_weight"] or "",
                    obj["total_machine_wastage"] or "",
                    obj["total_production_weight"] or "",
                    obj["total_weight24k"] or "",
                    obj["total_wastage_weight"] or "",
                    obj["total_weight24k"] or "",
                    obj["total_gold_received"] or "",
                    obj["balance_gold"],
                    obj["total_service_charges_amount"] or "",
                    obj["total_cash_received"] or "",
                    obj["balance_cash"]
                ]

        for col_idx, cell_value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_idx).value = cell_value

        row_num += 1

    # Add totals row
    row_num += 1
    if report_for == "Services":
        if report_type == "Detail":
            ws.cell(row=row_num, column=1, value="Grand Total:")
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
            ws.cell(row=row_num, column=7, value=totals['grand_total_estimate'])
            ws.cell(row=row_num, column=8, value=totals['grand_total_amount'])
        else:  # Summary
            ws.cell(row=row_num, column=1, value="Grand Total:")
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            ws.cell(row=row_num, column=3, value=totals['grand_total_estimate'])
            ws.cell(row=row_num, column=4, value=totals['grand_total_amount'])
    else:  # Casting
        if report_type == "Detail":
            ws.cell(row=row_num, column=1, value="Grand Total:")
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
            ws.cell(row=row_num, column=5, value=totals['grand_total_input'])
            ws.cell(row=row_num, column=6, value=totals['grand_total_output'])
            ws.cell(row=row_num, column=7, value=totals['grand_total_wastage'])
            ws.cell(row=row_num, column=10, value=totals['grand_total_production'])
            ws.cell(row=row_num, column=11, value=totals['grand_total_24k'])
            ws.cell(row=row_num, column=13, value=totals['grand_total_wastage_weight'])
            ws.cell(row=row_num, column=14, value=totals['grand_total_24k'])
            ws.cell(row=row_num, column=15, value=totals['grand_total_gold'])
            ws.cell(row=row_num, column=16, value=totals['grand_total_balance_gold'])
            ws.cell(row=row_num, column=18, value=totals['grand_total_service'])
            ws.cell(row=row_num, column=19, value=totals['grand_total_cash'])
            ws.cell(row=row_num, column=20, value=totals['grand_total_balance_cash'])
        else:  # Summary
            ws.cell(row=row_num, column=1, value="Grand Total:")
            ws.cell(row=row_num, column=2, value=totals['grand_total_input'])
            ws.cell(row=row_num, column=3, value=totals['grand_total_output'])
            ws.cell(row=row_num, column=4, value=totals['grand_total_wastage'])
            ws.cell(row=row_num, column=5, value=totals['grand_total_production'])
            ws.cell(row=row_num, column=6, value=totals['grand_total_24k'])
            ws.cell(row=row_num, column=7, value=totals['grand_total_wastage_weight'])
            ws.cell(row=row_num, column=8, value=totals['grand_total_24k'])
            ws.cell(row=row_num, column=9, value=totals['grand_total_gold'])
            ws.cell(row=row_num, column=10, value=totals['grand_total_balance_gold'])  # Balance Gold
            ws.cell(row=row_num, column=11, value=totals['grand_total_service'])  # Service Amount
            ws.cell(row=row_num, column=12, value=totals['grand_total_cash'])  # Received Amount
            ws.cell(row=row_num, column=13, value=totals['grand_total_balance_cash'])

            # Format totals row
    for row in ws.iter_rows(min_row=row_num, max_row=row_num):
        for cell in row:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{report_for.lower()}_{report_type.lower()}_{from_date}_{to_date}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
